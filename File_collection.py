##---------------
##---- collect all the files from a folder and subfolder location and group it in one folder .. for each path11,22,,etc
import os
from openpyxl import load_workbook

from pathlib import Path

#--- make sure you are in particular folder eg. LOC2

def collect(fpath):
    try:
        #os.chdir("C:/Users/stephen.arput/Documents/RESULTS/LOC222/LOC2")
        os.chdir(fpath)
        input_dir = Path.cwd()
        ### extacts all the files ending with path11.xlsx from subfolers also and saves it in Results_path11 folder
        for file in list(input_dir.rglob('*path11.xlsx')):
            wb=load_workbook(filename=file)
            output_dir = Path.cwd() / "Results_path11"
            output_dir.mkdir(exist_ok=True)
            wb.save(output_dir / file.name)
        print(" path11 files extraction completed")
        for file in list(input_dir.rglob('*path22.xlsx')):
            wb=load_workbook(filename=file)
            output_dir = Path.cwd() / "Results_path22"
            output_dir.mkdir(exist_ok=True)
            wb.save(output_dir / file.name)
        print(" path22 files extraction completed")
        for file in list(input_dir.rglob('*path33.xlsx')):
            wb=load_workbook(filename=file)
            output_dir = Path.cwd() / "Results_path33"
            output_dir.mkdir(exist_ok=True)
            wb.save(output_dir / file.name)
        print(" path33 files extraction completed")
        for file in list(input_dir.rglob('*path44.xlsx')):
            wb=load_workbook(filename=file)
            output_dir = Path.cwd() / "Results_path44"
            output_dir.mkdir(exist_ok=True)
            wb.save(output_dir / file.name)
        print(" path44 files extraction completed")
        for file in list(input_dir.rglob('*path55.xlsx')):
            wb=load_workbook(filename=file)
            output_dir = Path.cwd() / "Results_path55"
            output_dir.mkdir(exist_ok=True)
            wb.save(output_dir / file.name)
        print(" path55 files extraction completed")
        return ('Completed File collection in grouped folder')
    except Exception as e:
        return('The Exception message is:\n ', e)